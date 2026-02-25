"""
Job Application Tracker - Flask Web Application
================================================
A clean, full-featured job application tracking app.

Run Instructions:
    1. Install dependencies:   pip install -r requirements.txt
    2. Run the app:            python app.py
    3. Open browser at:        http://127.0.0.1:5000
"""

from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import date
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── App Setup ──────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.config['SECRET_KEY'] = 'jobtracker-secret-2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///jobs.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ── Database Model ─────────────────────────────────────────────────────────────
class JobApplication(db.Model):
    """Represents a single job application entry."""
    __tablename__ = 'job_applications'

    id           = db.Column(db.Integer, primary_key=True)
    company      = db.Column(db.String(120), nullable=False)
    role         = db.Column(db.String(120), nullable=False)
    date_applied = db.Column(db.Date, nullable=False, default=date.today)
    status       = db.Column(db.String(50), nullable=False, default='Pending')
    notes        = db.Column(db.Text, default='')

    def __repr__(self):
        return f'<JobApplication {self.company} – {self.role}>'


# ── Status Configuration ───────────────────────────────────────────────────────
STATUS_OPTIONS = ['Pending', 'Interview Scheduled', 'Selected', 'Rejected']

STATUS_COLORS = {
    'Pending':            {'bg': '#fff3cd', 'text': '#856404', 'badge': 'warning'},
    'Interview Scheduled':{'bg': '#cfe2ff', 'text': '#084298', 'badge': 'primary'},
    'Selected':           {'bg': '#d1e7dd', 'text': '#0a3622', 'badge': 'success'},
    'Rejected':           {'bg': '#f8d7da', 'text': '#842029', 'badge': 'danger'},
}


# ── Helper: Dashboard Counts ───────────────────────────────────────────────────
def get_summary():
    """Return aggregate counts for the dashboard."""
    total    = JobApplication.query.count()
    pending  = JobApplication.query.filter_by(status='Pending').count()
    rejected = JobApplication.query.filter_by(status='Rejected').count()
    selected = JobApplication.query.filter_by(status='Selected').count()
    interview = JobApplication.query.filter_by(status='Interview Scheduled').count()
    return dict(total=total, pending=pending, rejected=rejected,
                selected=selected, interview=interview)


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    """Main dashboard – list all applications with optional filter & search."""
    status_filter = request.args.get('status', '')
    search_query  = request.args.get('search', '').strip()

    query = JobApplication.query

    if status_filter and status_filter in STATUS_OPTIONS:
        query = query.filter_by(status=status_filter)

    if search_query:
        query = query.filter(JobApplication.company.ilike(f'%{search_query}%'))

    applications = query.order_by(JobApplication.date_applied.desc()).all()
    summary      = get_summary()

    return render_template(
        'index.html',
        applications=applications,
        summary=summary,
        status_options=STATUS_OPTIONS,
        status_colors=STATUS_COLORS,
        current_filter=status_filter,
        search_query=search_query,
    )


@app.route('/add', methods=['GET', 'POST'])
def add():
    """Add a new job application."""
    if request.method == 'POST':
        company      = request.form.get('company', '').strip()
        role         = request.form.get('role', '').strip()
        date_str     = request.form.get('date_applied', '')
        status       = request.form.get('status', 'Pending')
        notes        = request.form.get('notes', '').strip()

        # Basic validation
        if not company or not role or not date_str:
            flash('Company, Role, and Date Applied are required.', 'danger')
            return render_template('form.html', status_options=STATUS_OPTIONS,
                                   form_title='Add Application', action='add',
                                   today=date.today().isoformat())

        try:
            date_applied = date.fromisoformat(date_str)
        except ValueError:
            flash('Invalid date format.', 'danger')
            return redirect(url_for('add'))

        job = JobApplication(
            company=company,
            role=role,
            date_applied=date_applied,
            status=status,
            notes=notes,
        )
        db.session.add(job)
        db.session.commit()
        flash(f'Application to <strong>{company}</strong> added successfully!', 'success')
        return redirect(url_for('index'))

    return render_template('form.html', status_options=STATUS_OPTIONS,
                           form_title='Add Application', action='add',
                           today=date.today().isoformat(), job=None)


@app.route('/edit/<int:job_id>', methods=['GET', 'POST'])
def edit(job_id):
    """Edit an existing job application."""
    job = JobApplication.query.get_or_404(job_id)

    if request.method == 'POST':
        job.company      = request.form.get('company', '').strip()
        job.role         = request.form.get('role', '').strip()
        date_str         = request.form.get('date_applied', '')
        job.status       = request.form.get('status', 'Pending')
        job.notes        = request.form.get('notes', '').strip()

        if not job.company or not job.role or not date_str:
            flash('Company, Role, and Date Applied are required.', 'danger')
            return render_template('form.html', status_options=STATUS_OPTIONS,
                                   form_title='Edit Application', action=f'edit/{job_id}',
                                   job=job, today=date.today().isoformat())

        try:
            job.date_applied = date.fromisoformat(date_str)
        except ValueError:
            flash('Invalid date format.', 'danger')
            return redirect(url_for('edit', job_id=job_id))

        db.session.commit()
        flash(f'Application to <strong>{job.company}</strong> updated!', 'success')
        return redirect(url_for('index'))

    return render_template('form.html', status_options=STATUS_OPTIONS,
                           form_title='Edit Application', action=f'edit/{job_id}',
                           job=job, today=date.today().isoformat())


@app.route('/delete/<int:job_id>', methods=['POST'])
def delete(job_id):
    """Delete a job application (confirmed via modal on the frontend)."""
    job = JobApplication.query.get_or_404(job_id)
    company = job.company
    db.session.delete(job)
    db.session.commit()
    flash(f'Application to <strong>{company}</strong> deleted.', 'info')
    return redirect(url_for('index'))


@app.route('/export')
def export():
    """Export all applications to an Excel (.xlsx) file."""
    applications = JobApplication.query.order_by(JobApplication.date_applied.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Job Applications'

    # Header styling
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    headers = ['#', 'Company', 'Role', 'Date Applied', 'Status', 'Notes']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Status background colors for rows
    status_fill_map = {
        'Pending':             'FFF3CD',
        'Interview Scheduled': 'CFE2FF',
        'Selected':            'D1E7DD',
        'Rejected':            'F8D7DA',
    }

    # Data rows
    for row_num, job in enumerate(applications, 2):
        row_data = [
            row_num - 1,
            job.company,
            job.role,
            job.date_applied.strftime('%Y-%m-%d') if job.date_applied else '',
            job.status,
            job.notes or '',
        ]
        fill_color = status_fill_map.get(job.status, 'FFFFFF')
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    # Column widths
    col_widths = [5, 25, 25, 15, 22, 40]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 20

    # Save to buffer and send
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name='job_applications.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ── Entry Point ────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    with app.app_context():
        db.create_all()  # Create tables if they don't exist
    app.run(debug=True)
